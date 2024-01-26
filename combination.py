from itertools import combinations as com
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import numpy as np
import spacy

nlp = spacy.load("en_core_web_sm") # <--change-here
excel_file_name = "sample_book.xlsx" # <--change-here

def seed_combination(first_col, sec_col) :
    combi = []
    for first_cell in all [first_col][1:] :  
        for sec_cell in all[sec_col][1:] :
            if (first_cell.value is not None) and (sec_cell.value is not None) :
                combi.append(first_cell.value + " " + sec_cell.value)
            else : 
                break
    return combi

def create_comb_col_indices(indices):
    comb_indices = com(indices, 2)
    comb_indices_arr = [k for k in comb_indices]
    return comb_indices_arr


def create_sheet(sheet_name, first_col_ind, second_col_ind):
    ws = wb.create_sheet(str(sheet_name), 0)
    ws.title = str(sheet_name)
    combi = seed_combination(first_col_ind, second_col_ind)

    for row in combi:
        row_list = list()
        row_list.append(row)
        entities = nlp(row)
        for entity in entities.ents:
            row_list.append(entity.text) # append every entity in a new column
        ws.append(row_list)
    
wb = load_workbook(filename = excel_file_name)

titles = []
contentSeed_sheet = wb.get_sheet_by_name("Seeds")
last_col_letter = "A"
last_col_number = 0
for title_col in (contentSeed_sheet[1]):
    if title_col.value is not None:
        titles.append(title_col.value)
        last_col_letter = title_col.column_letter
        last_col_number = title_col.column
x = com(titles, 2)
col_names_combi = [i for i in x]

all = contentSeed_sheet["A:"+last_col_letter] 
col_indices = create_comb_col_indices(list(range(last_col_number)))
ind = 0
for col_name in col_names_combi :
    col_indice = col_indices[ind]
    sheet_name = col_name[0] + '_' + col_name[1]
    print("create sheet " + sheet_name)
    create_sheet(sheet_name, col_indice[0] , col_indice[1] )
    sheet_name = col_name[1] + '_' + col_name[0]
    print("create sheet " + sheet_name)
    create_sheet(sheet_name, col_indice[1] , col_indice[0] )
    ind = ind + 1
    wb.save(filename = excel_file_name)
